"""
kfnb_app/export/template_xlsx.py — 고객 전달 템플릿(.xlsx) 그대로 채우기.

업로드한 템플릿(예: Custom_F,B&T.xlsx)의 시트 구조·헤더·데이터사전을 보존하고,
우리 원천에서 만들 수 있는 시트를 채워 '완성된 엑셀'을 돌려준다.

채울 수 있는 시트:
  기본정보(회사정보): 공시명·브랜드명(회사_브랜드)·법인등록번호
  기본매출         : 거래일·회사명·브랜드명·채널·거래금액·거래수량·거래건수(·거래자수 미보유→공란)
채울 수 없는 시트(추가 데이터 필요): 인구통계(성별×연령)·재구매(코호트)·패널(DAU) → 헤더만 유지.

openpyxl 로 템플릿을 열어 해당 시트의 헤더 다음 행부터 기록한다. 헤더 위치는
키 헤더 토큰을 스캔해 자동 탐지(시트별 행 위치가 달라도 안전).
"""
from __future__ import annotations

import io
from pathlib import Path

import pandas as pd


# 시트 매칭(이름에 키워드 포함) → 채우기 스펙
def _find_header_row(ws, first_token: str, max_scan: int = 12):
    for r in range(1, min(ws.max_row or max_scan, max_scan) + 1):
        for c in range(1, 6):
            if str(ws.cell(r, c).value or "").strip() == first_token:
                return r, c
    return None, None


def _co_en_map(sm: pd.DataFrame) -> dict:
    if "company_en_official" not in sm.columns:
        return {}
    return (sm.dropna(subset=["company_kr"]).drop_duplicates("company_kr")
            .set_index("company_kr")["company_en_official"].astype(str).to_dict())


def _br_en_map(sm: pd.DataFrame) -> dict:
    if "brand_name_en" not in sm.columns:
        return {}
    return {(str(r["company_kr"]), str(r["brand_kr"])): str(r.get("brand_name_en") or "")
            for _, r in sm.drop_duplicates(["company_kr", "brand_kr"]).iterrows()}


def build_basic_info(sku_master: pd.DataFrame, *, english: bool = False) -> pd.DataFrame:
    """공시명·브랜드명(회사_브랜드, 회사_전체 포함)·법인등록번호. english=영문 표기."""
    sm = sku_master.copy()
    jr = "jurir_no" if "jurir_no" in sm.columns else None
    coen = _co_en_map(sm); bren = _br_en_map(sm)
    rows = []
    for co, g in sm.groupby("company_kr"):
        jur = (str(g[jr].dropna().iloc[0]) if jr and g[jr].notna().any() else "")
        cn = (coen.get(co) or co) if english else co
        tot = "Total" if english else "전체"
        rows.append({"공시명": cn, "브랜드명": f"{cn}_{tot}", "법인등록번호": jur})
        for br in sorted(set(str(b) for b in g["brand_kr"].dropna()
                             if str(b).strip() and str(b) not in ("(unknown)",))):
            bn = (bren.get((co, br)) or br) if english else br
            rows.append({"공시명": cn, "브랜드명": f"{cn}_{bn}", "법인등록번호": jur})
    return pd.DataFrame(rows, columns=["공시명", "브랜드명", "법인등록번호"])


def build_basic_sales(src, sku_master: pd.DataFrame, *, channel: str = "",
                      sector=None, with_company: bool = True,
                      english: bool = False) -> pd.DataFrame:
    """거래일×회사×브랜드 일별 집계 → 기본매출 레이아웃. english=회사/브랜드 영문."""
    d = src.daily_panel(sector).copy()
    coen = _co_en_map(sku_master); bren = _br_en_map(sku_master)
    d["거래일"] = d["date"].astype(str)
    if english:
        cn = d["company_kr"].map(lambda c: coen.get(str(c)) or str(c))
        bn = d.apply(lambda r: bren.get((str(r["company_kr"]), str(r["brand_kr"])))
                     or str(r["brand_kr"]), axis=1)
        d["회사명"] = cn
        d["브랜드명"] = cn.astype(str) + "_" + bn.astype(str)
    else:
        d["회사명"] = d["company_kr"]
        d["브랜드명"] = d["company_kr"].astype(str) + "_" + d["brand_kr"].astype(str)
    d["채널"] = channel or ""
    d["거래금액"] = d["sales_amt"]
    d["거래수량"] = d.get("sales_qty")
    d["거래건수"] = d.get("sales_cnt")
    d["거래자수"] = ""                              # 원천에 고유 거래자수 없음
    cols = (["거래일", "회사명", "브랜드명", "채널", "거래금액", "거래수량",
             "거래건수", "거래자수"] if with_company else
            ["거래일", "브랜드명", "채널", "거래금액", "거래수량", "거래건수", "거래자수"])
    return d[cols]


def _write_df(ws, df: pd.DataFrame, header_row: int, start_col: int = 1):
    """헤더 다음 행부터 df(헤더 순서대로) 기록. 기존 데이터행은 비우고 덮어씀."""
    # 헤더 순서 = 템플릿의 헤더셀 그대로 읽어 매핑
    headers = []
    c = start_col
    while True:
        v = ws.cell(header_row, c).value
        if v is None or str(v).strip() == "":
            break
        headers.append(str(v).strip()); c += 1
    # df 를 헤더 순서로 정렬(없는 컬럼은 공란)
    for i, h in enumerate(headers):
        col_vals = df[h] if h in df.columns else [""] * len(df)
        for j, val in enumerate(col_vals):
            ws.cell(header_row + 1 + j, start_col + i,
                    "" if pd.isna(val) else val)
    return len(df), headers


def fill_template(template_bytes: bytes, *, basic_info: pd.DataFrame,
                  basic_sales_prod: pd.DataFrame,
                  basic_sales_bt: pd.DataFrame | None = None) -> tuple[bytes, dict]:
    """템플릿 bytes → 채운 xlsx bytes + 채움 리포트."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    report = {}
    for ws in wb.worksheets:
        name = ws.title
        if "data_dictionary" in name:
            continue
        if "기본정보" in name:
            hr, hc = _find_header_row(ws, "공시명")
            if hr:
                n, _ = _write_df(ws, basic_info, hr, hc)
                report[name] = f"기본정보 {n}행"
        elif "기본매출" in name:
            df = basic_sales_bt if ("백테스트" in name and basic_sales_bt is not None) \
                else basic_sales_prod
            hr, hc = _find_header_row(ws, "거래일")
            if hr:
                n, _ = _write_df(ws, df, hr, hc)
                report[name] = f"기본매출 {n}행"
        else:
            report[name] = "헤더 유지(추가 데이터 필요)"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), report
