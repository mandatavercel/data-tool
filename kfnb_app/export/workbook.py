"""
kfnb_app/export.py — ⑥ 최종 데이터 상품(xlsx) 생성.

파이프라인 산출물(정규화·태깅·매핑된 SKU 마스터, 월별 패널, 연별 모멘텀,
프로파일)을 받아 고객에게 건넬 다중시트 xlsx 를 만든다. 모든 계산 컬럼은
IFERROR 로 가드된 Excel 수식으로 작성해 수식 오류 0 을 보장한다.
openpyxl 만 사용 (LibreOffice 불필요 — Excel 이 열 때 수식 계산).
"""
from __future__ import annotations

from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
HEAD_FILL = PatternFill("solid", fgColor="1F3864")
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
_thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _f(**k) -> Font:
    return Font(name=FONT, **k)


def _style_header(ws, row: int, ncol: int):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL
        cell.font = _f(color="FFFFFF", bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def _write_table(ws, df: pd.DataFrame, headers: list[str], start: int = 1,
                 num_cols: Optional[dict] = None, text_cols: Optional[set] = None):
    num_cols = num_cols or {}
    text_cols = text_cols or set()
    for j, h in enumerate(headers, 1):
        ws.cell(row=start, column=j, value=h)
    _style_header(ws, start, len(headers))
    for i, (_, r) in enumerate(df.iterrows(), start + 1):
        for j, col in enumerate(df.columns, 1):
            v = r[col]
            cell = ws.cell(row=i, column=j,
                           value=(None if pd.isna(v) else v))
            cell.font = _f(size=10)
            cell.border = BORDER
            if col in num_cols:
                cell.number_format = num_cols[col]
            if col in text_cols:
                cell.number_format = "@"
    return start + len(df)


# 고객 유형 → 포함 시트 (README 는 항상 포함)
PRODUCT_SHEETS: dict[str, set] = {
    "quant":       {"sku", "panel"},          # PIT 패널 + 티커 매핑
    "fundamental": {"sku", "momentum"},       # SKU 트래커 + 모멘텀
    "vendor":      {"sku", "source_request"}, # 정규화 raw feed + 사전
}


def build_workbook(out_path: str | Path, *, profile: dict, sku_master: pd.DataFrame,
                   monthly_panel: pd.DataFrame, annual_company: pd.DataFrame,
                   brand_trend: Optional[pd.DataFrame] = None,
                   mapping_report: Optional[dict] = None,
                   source_name: str = "(원천 미상)",
                   sector_label: str = "K-F&B",
                   focus_brand: str = "",
                   products: Optional[list[str]] = None,
                   brand_master: Optional[pd.DataFrame] = None,
                   use_cases: Optional[pd.DataFrame] = None,
                   analysis_cols: Optional[list[str]] = None,
                   id_cols: Optional[list[str]] = None) -> dict:
    """xlsx 생성. products 로 고객유형별 시트 구성. 반환: {path, sheets, customers}."""
    wb = Workbook()
    mapping_report = mapping_report or {}

    # 포함할 시트 집합 결정 (products 미지정 = 전체)
    products = products or list(PRODUCT_SHEETS.keys())
    want: set = set()
    for p in products:
        want |= PRODUCT_SHEETS.get(p, set())
    if not want:
        want = {"sku", "panel", "momentum", "source_request"}

    # ── ① README & Health ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "① README & Health"
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"{sector_label} POS 데이터 상품 — Mandata"
    ws["A1"].font = _f(bold=True, size=16, color="1F3864")
    ws["A2"] = f"원천: {source_name}  |  생성: K-F&B Product Agent (kfnb_app)"
    ws["A2"].font = _f(size=10, italic=True, color="595959")

    s = profile["summary"]
    q = profile["quality"]
    rows = [
        ("", ""),
        ("데이터 헬스", ""),
        ("항목", "값"),
        ("기간", s["period"]),
        ("총 레코드", f"{s['rows']:,} 행"),
        ("커버리지", f"회사 {s['companies']} · 브랜드 {s['brands']} · "
                     f"SKU {s['skus']} · 지역 {s['regions']} · {s['days']}일"),
        ("섹터(cat_l1)", ", ".join(profile.get("sectors", [])) or "?"),
        ("핵심 결측", f"회사 null {q['null_company']:,} · SKU null {q['null_sku']:,}"),
        ("음수/0 매출", f"{q['nonpos_amt']:,}행 ({q['nonpos_pct']:.2f}%)"),
        ("바코드(EAN-13)", f"{q['barcode_ok_pct']:.1f}% 13자리"),
        ("상장사", ", ".join(mapping_report.get("listed", [])) or "?"),
        ("비상장사", ", ".join(mapping_report.get("private", [])) or "-"),
        ("", ""),
        ("시트 안내", ""),
        ("② SKU_Master", "정규화·투자태그·티커 매핑된 SKU 마스터 (재사용 핵심 자산)"),
        ("③ Monthly_Panel", "월별 회사·브랜드 매출 패널 (백테스트용)"),
        ("④ Company_Momentum", "연도별 회사 매출·YoY" +
                                (f" + {focus_brand} 추세" if focus_brand else "")),
        ("⑤ Source_Request", "원천사 추가요청 필드 (투자등급 승격 조건)"),
    ]
    r = 4
    for a, b in rows:
        ws.cell(row=r, column=1, value=a)
        ws.cell(row=r, column=2, value=b)
        if a in ("데이터 헬스", "시트 안내"):
            ws.cell(row=r, column=1).font = _f(bold=True, size=13, color="1F3864")
        elif a == "항목":
            for c in (1, 2):
                ws.cell(row=r, column=c).fill = SUB_FILL
                ws.cell(row=r, column=c).font = _f(bold=True)
        else:
            ws.cell(row=r, column=1).font = _f(bold=True, size=10)
            ws.cell(row=r, column=2).font = _f(size=10)
            ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 92

    # ── ② SKU_Master (선택 분석 컬럼 포함) ──────────────────────────────
    from kfnb_app.mapping import mastering as _mast
    ws = wb.create_sheet("② SKU_Master")
    sub = _mast.build_sku_master_file(sku_master, analysis_cols, id_cols)
    cols = list(sub.columns)
    _write_table(ws, sub, cols,
                 num_cols={"sales_amt": "#,##0", "sales_qty": "#,##0",
                           "asp_won": "#,##0"},
                 text_cols={"barcode"})
    ws.freeze_panes = "A2"
    for j, c in enumerate(cols, 1):
        w = 24 if "name" in c else (16 if c in ("sku_id", "brand_id",
            "company_name_en", "company_kr") else 11)
        ws.column_dimensions[get_column_letter(j)].width = w

    # ── ③ Monthly_Panel ────────────────────────────────────────────────
    ws = wb.create_sheet("③ Monthly_Panel")
    pcols = ["ym", "company_kr", "bbg_ticker", "brand_kr",
             "sales_amt", "sales_qty", "receipts"]
    pcols = [c for c in pcols if c in monthly_panel.columns]
    pheaders = ["YYYYMM", "회사", "Ticker", "브랜드", "Sales Amt(₩)",
                "Sales Qty", "Receipts"][:len(pcols)]
    mp = monthly_panel[pcols].copy()
    if "ym" in mp:
        mp["ym"] = mp["ym"].astype(int).astype(str)
    end = _write_table(ws, mp, pheaders,
                       num_cols={"sales_amt": "#,##0", "sales_qty": "#,##0",
                                 "receipts": "#,##0"})
    if "sales_amt" in pcols and "sales_qty" in pcols:
        ac = get_column_letter(pcols.index("sales_amt") + 1)
        qc = get_column_letter(pcols.index("sales_qty") + 1)
        asp_col = len(pcols) + 1
        ws.cell(row=1, column=asp_col, value="ASP(₩)")
        _style_header(ws, 1, asp_col)
        for i in range(2, end):
            cell = ws.cell(row=i, column=asp_col,
                           value=f"=IFERROR({ac}{i}/{qc}{i},0)")
            cell.number_format = "#,##0"
            cell.font = _f(size=10)
            cell.border = BORDER
    ws.freeze_panes = "A2"
    for j, w in enumerate([10, 12, 11, 16, 15, 12, 10, 10][:len(pcols) + 1], 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # ── ④ Company_Momentum ─────────────────────────────────────────────
    ws = wb.create_sheet("④ Company_Momentum")
    ws["A1"] = "연도별 회사 매출 + YoY%"
    ws["A1"].font = _f(bold=True, size=13, color="1F3864")
    ac = annual_company.copy()
    accols = ["yr", "company_kr", "bbg_ticker", "sales_amt", "sales_qty", "yoy_pct"]
    accols = [c for c in accols if c in ac.columns]
    acheaders = ["연도", "회사", "Ticker", "Sales Amt(₩)", "Sales Qty", "YoY%"][:len(accols)]
    if "yr" in ac:
        ac["yr"] = ac["yr"].astype(int).astype(str)
    _write_table(ws, ac[accols], acheaders, start=3,
                 num_cols={"sales_amt": "#,##0", "sales_qty": "#,##0",
                           "yoy_pct": "0.0"})
    ws.cell(row=3 + len(ac) + 2, column=1,
            value="※ 부분 연도(예: 당해 1~3월)는 연환산 아님 — 절대값 비교 주의").font = \
        _f(size=9, italic=True, color="595959")

    if brand_trend is not None and not brand_trend.empty:
        b0 = 3 + len(ac) + 4
        ws.cell(row=b0, column=1, value=f"{focus_brand} — 브랜드 모멘텀").font = \
            _f(bold=True, size=13, color="C00000")
        bt = brand_trend.copy()
        btcols = ["yr", "sales_amt", "sales_qty", "asp_won", "yoy_pct"]
        btcols = [c for c in btcols if c in bt.columns]
        btheaders = ["연도", "Sales Amt(₩)", "Qty", "ASP(₩)", "YoY%"][:len(btcols)]
        if "yr" in bt:
            bt["yr"] = bt["yr"].astype(int).astype(str)
        _write_table(ws, bt[btcols], btheaders, start=b0 + 1,
                     num_cols={"sales_amt": "#,##0", "sales_qty": "#,##0",
                               "asp_won": "#,##0", "yoy_pct": "0.0"})
    ws.column_dimensions["A"].width = 22
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 15

    # ── ⑤ Source_Request ───────────────────────────────────────────────
    ws = wb.create_sheet("⑤ Source_Request")
    ws["A1"] = "원천사 추가요청 필드 — 투자등급 승격 조건"
    ws["A1"].font = _f(bold=True, size=13, color="1F3864")
    brows = [
        ["필드", "왜 필요한가 (투자 관점)", "우선순위"],
        ["promo_flag / event_type", "행사(2+1·할인) 분리 → 진성수요 vs 프로모션", "★★★ 필수"],
        ["regular_price / actual_price", "정가 대비 실판매가 → 가격탄력성·마진", "★★★ 필수"],
        ["store_count_selling", "판매 점포수 → 동일점(SSS) vs 입점확대 분리", "★★★ 필수"],
        ["basket / receipt id", "동시구매 분석 → 소비자 행동", "★★ 권장"],
        ["age/gender bucket", "인구통계 → 트렌드 민감도", "★★ 권장"],
        ["return/cancel flag", "음수 매출 정의 명확화", "★ 참고"],
    ]
    for i, row in enumerate(brows, 3):
        for j, v in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=v)
            if i == 3:
                cell.fill = HEAD_FILL
                cell.font = _f(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
            else:
                cell.font = _f(size=10)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = BORDER
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 56
    ws.column_dimensions["C"].width = 12

    # ── ⑥ Brand_Master (딕셔너리 — 항상 포함) ──────────────────────────
    if brand_master is not None and not brand_master.empty:
        ws = wb.create_sheet("⑥ Brand_Master")
        bcols = ["brand_id", "company_name_en", "brand_name_ko", "brand_name_en",
                 "brand_aliases", "skus", "sales_amt", "mapping_status",
                 "mapping_confidence"]
        bcols = [c for c in bcols if c in brand_master.columns]
        bheaders = ["Brand ID", "Company(EN)", "브랜드(KR)", "브랜드(EN)",
                    "Aliases", "SKUs", "Sales(₩)", "Status", "Conf."][:len(bcols)]
        _write_table(ws, brand_master[bcols], bheaders,
                     num_cols={"sales_amt": "#,##0", "skus": "#,##0"})
        ws.freeze_panes = "A2"
        for j, w in enumerate([26, 22, 14, 16, 40, 7, 15, 12, 8][:len(bcols)], 1):
            ws.column_dimensions[get_column_letter(j)].width = w

    # ── ⑦ Use_Cases (전수 발굴 시그널 — 항상 포함) ─────────────────────
    if use_cases is not None and not use_cases.empty:
        ws = wb.create_sheet("⑦ Use_Cases")
        ucols = ["rank", "usecase_type", "entity_kr", "entity_en", "ticker",
                 "isin", "metric", "value", "window", "direction",
                 "confidence", "thesis_en"]
        ucols = [c for c in ucols if c in use_cases.columns]
        uheaders = ["Rank", "Type", "Entity(KR)", "Entity(EN)", "Ticker",
                    "ISIN", "Metric", "Value", "Window", "Dir",
                    "Conf.", "Thesis (EN)"][:len(ucols)]
        _write_table(ws, use_cases[ucols], uheaders,
                     num_cols={"value": "#,##0.0", "rank": "0"})
        ws.freeze_panes = "A2"
        for j, w in enumerate([6, 13, 16, 18, 11, 14, 18, 11, 13, 6, 7, 52][:len(ucols)], 1):
            ws.column_dimensions[get_column_letter(j)].width = w

    # ── 고객유형에 맞지 않는 시트 제거 (README 는 항상 유지) ──────────────
    sheet_key = {
        "② SKU_Master": "sku",
        "③ Monthly_Panel": "panel",
        "④ Company_Momentum": "momentum",
        "⑤ Source_Request": "source_request",
    }
    for name, key in sheet_key.items():
        if key not in want and name in wb.sheetnames:
            wb.remove(wb[name])

    out_path = str(out_path)
    wb.save(out_path)
    return {"path": out_path, "sheets": len(wb.sheetnames),
            "customers": products}
