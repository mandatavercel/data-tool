"""
mapping_app/master_builder.py — POS 회사 마스터 시트 빌더.

새 모드 (회사 마스터 빌더) 의 핵심 모듈.
입력: 회사 한글 이름 리스트
출력: 4-시트 xlsx (INFORMATION_PR / INFORMATION_BT / LIST_PR / LIST_BT)

LIST 시트 컬럼 (첨부 파일 레이아웃 그대로):
  gics_industry_code, gics_industry, gics_sub_industry_code, gics_sub_industry,
  company_name_en, korea_company_number, mandata_brand_name, mandata_brand_code,
  listing_status, security_code

INFORMATION 시트: header(행5) + 데이터(행6+) — POS Data Introduction 설명 페이지.

흐름:
  M1. 회사 이름 입력
  M2. DART 매칭 (corp_code, jurir_no, eng_name, stock_code, induty_code, corp_cls)
  M3. ISIN 산출 (security_code, listing_status)
  M4. GICS 매핑 (induty_code → gics_4컬럼, 수동 override 가능)
  M5. INFORMATION 메타 form
  M6. 4-시트 xlsx 생성 + 다운로드
"""
from __future__ import annotations

import io
from datetime import datetime
from pathlib import Path

import pandas as pd


# ══════════════════════════════════════════════════════════════════════════════
# KSIC (DART induty_code) → GICS 매핑 테이블
# ══════════════════════════════════════════════════════════════════════════════
# DART API 의 induty_code 는 한국 표준산업분류(KSIC) 6자리.
# GICS sub_industry_code 는 8자리. 정확한 1:1 매핑은 어려우니
# 흔한 업종 위주로 매핑. 없으면 사용자가 수동 입력.

# KSIC 앞 2자리 (대분류) 또는 앞 4자리 (중분류) 기준으로 매핑.
# (gics_industry_code, gics_industry, gics_sub_industry_code, gics_sub_industry)
GICS_BY_KSIC: dict[str, tuple[int, str, int, str]] = {
    # ── 식품 가공·제조 ─────────────────────────────────────────────────────
    "107":   (302020, "Food Products",          30202030, "Packaged Foods & Meats"),
    "1071":  (302020, "Food Products",          30202030, "Packaged Foods & Meats"),
    "1072":  (302020, "Food Products",          30202030, "Packaged Foods & Meats"),
    "1073":  (302020, "Food Products",          30202030, "Packaged Foods & Meats"),
    "1074":  (302020, "Food Products",          30202030, "Packaged Foods & Meats"),
    "1075":  (302020, "Food Products",          30202030, "Packaged Foods & Meats"),
    "1076":  (302020, "Food Products",          30202030, "Packaged Foods & Meats"),
    "1077":  (302020, "Food Products",          30202030, "Packaged Foods & Meats"),
    "1079":  (302020, "Food Products",          30202030, "Packaged Foods & Meats"),
    "108":   (302020, "Food Products",          30202030, "Packaged Foods & Meats"),
    # ── 음료 ──────────────────────────────────────────────────────────────
    "111":   (302010, "Beverages",              30201030, "Soft Drinks"),
    "1111":  (302010, "Beverages",              30201010, "Brewers"),
    "1112":  (302010, "Beverages",              30201020, "Distillers & Vintners"),
    "1113":  (302010, "Beverages",              30201020, "Distillers & Vintners"),
    "1114":  (302010, "Beverages",              30201030, "Soft Drinks"),
    # ── 담배 ──────────────────────────────────────────────────────────────
    "120":   (302030, "Tobacco",                30203010, "Tobacco"),
    # ── 화장품·생활용품 (Household & Personal Products) ────────────────────
    "204":   (303020, "Personal Products",      30302010, "Personal Products"),
    "2043":  (303020, "Personal Products",      30302010, "Personal Products"),
    # ── 화학 (Diversified Chemicals 등) ────────────────────────────────────
    "201":   (151010, "Chemicals",              15101020, "Diversified Chemicals"),
    "202":   (151010, "Chemicals",              15101030, "Specialty Chemicals"),
    "203":   (151010, "Chemicals",              15101030, "Specialty Chemicals"),
    # ── 의약품 ────────────────────────────────────────────────────────────
    "21":    (352010, "Pharmaceuticals",        35201010, "Pharmaceuticals"),
    "212":   (352010, "Pharmaceuticals",        35201010, "Pharmaceuticals"),
    # ── 식료품 도매·소매 (Food & Staples Retailing) ───────────────────────
    "463":   (301010, "Food & Staples Retailing", 30101020, "Food Distributors"),
    "472":   (301010, "Food & Staples Retailing", 30101030, "Food Retail"),
    "4711":  (301010, "Food & Staples Retailing", 30101010, "Drug Retail"),
    "4712":  (301010, "Food & Staples Retailing", 30101040, "Hypermarkets & Super Centers"),
    # ── 반도체 ────────────────────────────────────────────────────────────
    "261":   (453010, "Semiconductors & Semiconductor Equipment", 45301020, "Semiconductors"),
    # ── 자동차 ────────────────────────────────────────────────────────────
    "30":    (251020, "Automobiles",            25102010, "Automobile Manufacturers"),
    # ── 은행 / 금융 ───────────────────────────────────────────────────────
    "641":   (401010, "Banks",                  40101010, "Diversified Banks"),
    "642":   (402010, "Diversified Financial Services", 40201040, "Multi-Sector Holdings"),
}


def map_ksic_to_gics(ksic_code: str | int | None) -> tuple[int, str, int, str] | None:
    """KSIC induty_code (6자리) → GICS 4-tuple. 매핑 없으면 None.

    1) 사용자 누적 캐시 csv (정확 매칭 6자리 우선)
    2) builtin GICS_BY_KSIC (긴 prefix 부터)
    """
    if not ksic_code:
        return None
    s = str(ksic_code).strip()
    if not s.isdigit():
        return None
    # ① 캐시: 정확 6자리 매칭 우선
    cache = _load_ksic_gics_cache()
    if s in cache:
        return cache[s]
    # ② builtin: 긴 prefix 부터
    for n in (6, 4, 3, 2):
        prefix = s[:n]
        if prefix in GICS_BY_KSIC:
            return GICS_BY_KSIC[prefix]
    # ③ 캐시 prefix fallback (캐시에 4자리 등록된 게 있을 수 있음)
    for n in (4, 3, 2):
        prefix = s[:n]
        if prefix in cache:
            return cache[prefix]
    return None


# ══════════════════════════════════════════════════════════════════════════════
# KSIC → GICS 누적 캐시 (csv) — LLM 추론 결과를 저장해 다음 빌드에 자동 적용
# ══════════════════════════════════════════════════════════════════════════════

_CACHE_DIR  = Path(__file__).resolve().parent / "data"
_CACHE_PATH = _CACHE_DIR / "ksic_gics_cache.csv"
_CACHE_COLS = [
    "ksic_code", "gics_industry_code", "gics_industry",
    "gics_sub_industry_code", "gics_sub_industry",
    "source", "confidence", "updated_at",
]
# 프로세스 캐시 — 한 세션 동안 csv 를 매번 읽지 않도록
_CACHE_MEM: dict[str, tuple[int, str, int, str]] | None = None


def _load_ksic_gics_cache() -> dict[str, tuple[int, str, int, str]]:
    """csv → {ksic_code: (gics_ind_code, gics_ind, gics_sub_code, gics_sub)}."""
    global _CACHE_MEM
    if _CACHE_MEM is not None:
        return _CACHE_MEM
    if not _CACHE_PATH.exists():
        _CACHE_MEM = {}
        return _CACHE_MEM
    try:
        df = pd.read_csv(_CACHE_PATH, dtype={"ksic_code": str})
        out: dict[str, tuple[int, str, int, str]] = {}
        for _, r in df.iterrows():
            k = str(r["ksic_code"]).strip()
            try:
                out[k] = (
                    int(r["gics_industry_code"]),
                    str(r["gics_industry"]),
                    int(r["gics_sub_industry_code"]),
                    str(r["gics_sub_industry"]),
                )
            except (ValueError, TypeError):
                continue
        _CACHE_MEM = out
        return out
    except Exception:
        _CACHE_MEM = {}
        return {}


def save_ksic_gics_cache_rows(rows: list[dict]) -> int:
    """LLM 추론·수동 확정 매핑을 캐시 csv 에 추가/갱신.

    rows: [{ksic_code, gics_industry_code, gics_industry,
            gics_sub_industry_code, gics_sub_industry, source, confidence}, ...]
    같은 ksic_code 있으면 source/confidence 가 더 높은 쪽으로 덮어씀.
    Returns: 새로 추가/갱신된 row 수.
    """
    if not rows:
        return 0
    _CACHE_DIR.mkdir(parents=True, exist_ok=True)
    # 기존 로드
    existing: dict[str, dict] = {}
    if _CACHE_PATH.exists():
        try:
            df = pd.read_csv(_CACHE_PATH, dtype={"ksic_code": str})
            for _, r in df.iterrows():
                existing[str(r["ksic_code"]).strip()] = r.to_dict()
        except Exception:
            existing = {}

    _SRC_PRIO = {"manual": 3, "builtin": 2, "llm": 1, "": 0}
    n_changed = 0
    for r in rows:
        k = str(r.get("ksic_code", "")).strip()
        gic = int(r.get("gics_industry_code", 0) or 0)
        sic = int(r.get("gics_sub_industry_code", 0) or 0)
        if not k or not gic or not sic:
            continue
        new_row = {
            "ksic_code":              k,
            "gics_industry_code":     gic,
            "gics_industry":          str(r.get("gics_industry", "")),
            "gics_sub_industry_code": sic,
            "gics_sub_industry":      str(r.get("gics_sub_industry", "")),
            "source":                 str(r.get("source", "llm")),
            "confidence":             float(r.get("confidence", 0.7) or 0.0),
            "updated_at":             datetime.now().isoformat(timespec="seconds"),
        }
        if k in existing:
            prev = existing[k]
            new_prio = _SRC_PRIO.get(new_row["source"], 0)
            old_prio = _SRC_PRIO.get(str(prev.get("source", "")), 0)
            if new_prio < old_prio:
                continue   # 옛 항목이 더 우선순위 높음 (예: manual > llm)
            if new_prio == old_prio and new_row["confidence"] <= float(prev.get("confidence", 0)):
                continue
        existing[k] = new_row
        n_changed += 1

    # 저장
    out_df = pd.DataFrame(list(existing.values()), columns=_CACHE_COLS)
    out_df.to_csv(_CACHE_PATH, index=False)
    # 프로세스 캐시 무효화 — 다음 호출에 다시 읽음
    global _CACHE_MEM
    _CACHE_MEM = None
    return n_changed


def reload_ksic_gics_cache() -> None:
    """캐시 무효화 — 외부에서 csv 수동 편집한 뒤 호출."""
    global _CACHE_MEM
    _CACHE_MEM = None


def llm_classify_gics(
    items: list[dict],
    api_key: str | None,
) -> list[dict]:
    """누락 GICS 행에 대해 LLM 추론. items: [{ksic_code, name_kr, name_en, stock_code}].

    Returns: 같은 순서의 GICS 4-tuple dict 리스트. 빈 dict 면 실패.
    """
    if not items or not api_key:
        return [{} for _ in items]
    try:
        from modules.mapping.translation import llm as _llm
        return _llm.llm_classify_gics_batch(items, api_key, chunk_size=40, max_workers=4)
    except Exception:
        return [{} for _ in items]


# ══════════════════════════════════════════════════════════════════════════════
# 자회사 → 모회사 매핑 (비상장/Delisted 처리)
# ══════════════════════════════════════════════════════════════════════════════
# 자회사 매핑은 보통 안정적이므로 (동원F&B → 동원산업) 캐시로 누적.
# UI 흐름: M3 ISIN 단계에서 비상장 회사들 표시 → LLM 추정/수동 입력 → 캐시 저장

_SUB_CACHE_PATH = _CACHE_DIR / "subsidiary_cache.csv"
_SUB_CACHE_COLS = [
    "company_kr", "parent_kr", "parent_en",
    "parent_stock_code", "parent_isin",
    "status_kind",                     # "delisted" | "subsidiary" | "international"
    "source", "confidence", "updated_at",
]
_SUB_CACHE_MEM: dict[str, dict] | None = None


def load_subsidiary_cache() -> dict[str, dict]:
    """csv → {company_kr: {parent_kr, parent_en, parent_stock_code, parent_isin, status_kind}}."""
    global _SUB_CACHE_MEM
    if _SUB_CACHE_MEM is not None:
        return _SUB_CACHE_MEM
    if not _SUB_CACHE_PATH.exists():
        _SUB_CACHE_MEM = {}
        return _SUB_CACHE_MEM
    try:
        df = pd.read_csv(_SUB_CACHE_PATH, dtype=str).fillna("")
        out = {}
        for _, r in df.iterrows():
            k = str(r["company_kr"]).strip()
            if k:
                out[k] = {
                    "parent_kr":          r["parent_kr"],
                    "parent_en":          r["parent_en"],
                    "parent_stock_code":  r["parent_stock_code"],
                    "parent_isin":        r["parent_isin"],
                    "status_kind":        r["status_kind"] or "subsidiary",
                }
        _SUB_CACHE_MEM = out
        return out
    except Exception:
        _SUB_CACHE_MEM = {}
        return {}


def save_subsidiary_cache_rows(rows: list[dict]) -> int:
    """자회사 매핑 누적 저장. rows: [{company_kr, parent_kr, parent_en, parent_stock_code,
                                       parent_isin, status_kind, source, confidence}, ...]"""
    if not rows:
        return 0
    _CACHE_DIR.mkdir(parents=True, exist_ok=True)
    existing: dict[str, dict] = {}
    if _SUB_CACHE_PATH.exists():
        try:
            df = pd.read_csv(_SUB_CACHE_PATH, dtype=str).fillna("")
            for _, r in df.iterrows():
                existing[str(r["company_kr"]).strip()] = r.to_dict()
        except Exception:
            existing = {}

    _SRC_PRIO = {"manual": 3, "builtin": 2, "llm": 1, "": 0}
    n_changed = 0
    for r in rows:
        k = str(r.get("company_kr", "")).strip()
        if not k:
            continue
        new_row = {
            "company_kr":         k,
            "parent_kr":          str(r.get("parent_kr", "")),
            "parent_en":          str(r.get("parent_en", "")),
            "parent_stock_code":  str(r.get("parent_stock_code", "")),
            "parent_isin":        str(r.get("parent_isin", "")),
            "status_kind":        str(r.get("status_kind", "subsidiary")),
            "source":             str(r.get("source", "llm")),
            "confidence":         str(r.get("confidence", 0.8) or 0.0),
            "updated_at":         datetime.now().isoformat(timespec="seconds"),
        }
        if k in existing:
            prev = existing[k]
            new_prio = _SRC_PRIO.get(new_row["source"], 0)
            old_prio = _SRC_PRIO.get(str(prev.get("source", "")), 0)
            if new_prio < old_prio:
                continue
        existing[k] = new_row
        n_changed += 1

    out_df = pd.DataFrame(list(existing.values()), columns=_SUB_CACHE_COLS)
    out_df.to_csv(_SUB_CACHE_PATH, index=False)
    global _SUB_CACHE_MEM
    _SUB_CACHE_MEM = None
    return n_changed


def format_subsidiary_status(parent_kr: str, parent_isin: str, status_kind: str = "subsidiary") -> str:
    """모회사 정보 → listing_status 텍스트 (원본 POS 파일 패턴과 동일)."""
    parent_kr = (parent_kr or "").strip()
    parent_isin = (parent_isin or "").strip()
    if not parent_kr:
        return "Not listed"
    if status_kind == "delisted":
        prefix = "Delisted"
    elif status_kind == "international":
        prefix = "Internationally Listed"
    else:
        prefix = "Subsidiary"
    if parent_isin:
        return f"{prefix} (subsidiary of {parent_kr} - {parent_isin})"
    return f"{prefix} (subsidiary of {parent_kr})"


def apply_subsidiary_to_row(row: dict, parent_info: dict | None) -> dict:
    """build_company_list_row 결과에 모회사 정보 덮어쓰기.

    parent_info 가 있고 parent_isin 도 있으면:
      security_code → parent_isin
      listing_status → 'Subsidiary of {parent_kr} ({parent_isin})'
    """
    if not parent_info:
        return row
    parent_kr   = (parent_info.get("parent_kr") or "").strip()
    parent_isin = (parent_info.get("parent_isin") or "").strip()
    if not parent_kr:
        return row
    new_row = dict(row)
    new_row["listing_status"] = format_subsidiary_status(
        parent_kr, parent_isin,
        status_kind=parent_info.get("status_kind", "subsidiary"),
    )
    if parent_isin:
        new_row["security_code"] = parent_isin
    return new_row


def resolve_shareholder_from_dart_master(
    shareholder_name_kr: str,
    dart_master: "pd.DataFrame",
) -> dict:
    """DART 최대주주 한글명을 dart_master 에서 찾아 영문/상장정보/ISIN 산출.

    Returns:
        {largest_shareholder_company_name_en, largest_shareholder_listing_status,
         largest_shareholder_security_code} — 매칭 실패시 빈 dict.
    """
    if not shareholder_name_kr or dart_master is None or dart_master.empty:
        return {}

    name = shareholder_name_kr.strip()

    # 주요 기관 휴리스틱 (DART 마스터에 없거나 별도 처리)
    if "국민연금" in name or "NPS" in name.upper():
        return {
            "largest_shareholder_company_name_en": "NATIONAL_PENSION_SERVICE",
            "largest_shareholder_listing_status":  "N/A",
            "largest_shareholder_security_code":   "N/A",
            "_source":      "Institution (NPS)",
            "_master_match": "국민연금공단",
        }

    # 1순위 — DART 마스터 정확 일치
    match_row = None
    match_kind = ""   # 'exact' / 'prefix_suffix' / ''
    matched = dart_master[dart_master["corp_name"].astype(str).str.strip() == name]
    if not matched.empty:
        match_row = matched.iloc[0]
        match_kind = "exact"
    elif len(name) >= 3:
        # 부분 매칭 (엄격) — name 이 corp_name 의 prefix/suffix 일 때만.
        # substring(중간) 매칭은 '이' / '세' 같은 짧은 한글이 임의 회사에 걸리는 문제
        # 때문에 사용 안 함. 길이비도 50% 이상이어야.
        def _prefix_or_suffix(cn: str) -> bool:
            cn = str(cn or "").strip()
            if not cn or len(cn) < 3:
                return False
            if cn == name:
                return True
            if cn.startswith(name) or cn.endswith(name):
                # 길이비 0.5 이상 (예: '농심'→'농심홀딩스' 2/5=0.4 통과 X,
                # '농심홀딩스'→'농심홀딩스' OK)
                return min(len(name), len(cn)) / max(len(name), len(cn)) >= 0.5
            return False
        cands = dart_master[
            dart_master["corp_name"].astype(str).apply(_prefix_or_suffix)
        ]
        if not cands.empty:
            listed = cands[cands["stock_code"].astype(str).str.strip() != ""]
            match_row = (listed.iloc[0] if not listed.empty else cands.iloc[0])
            match_kind = "prefix_suffix"

    # 매칭 실패 시에만 자연인 휴리스틱 — 한글 2~4자 + 회사 접미사 없음
    if match_row is None:
        _COMPANY_SUFFIX = ("산업","기업","그룹","홀딩스","제약","화학","식품","전자",
                           "통신","은행","증권","카드","건설","에너지","바이오",
                           "공사","공단","재단","협회","유한회사","주식회사")
        is_company_suffix = any(s in name for s in _COMPANY_SUFFIX)
        is_short_korean = (
            2 <= len(name) <= 4
            and all('가' <= c <= '힣' for c in name)
            and not is_company_suffix
        )
        if is_short_korean:
            return {
                "largest_shareholder_company_name_en": name,
                "largest_shareholder_listing_status":  "N/A (Individual)",
                "largest_shareholder_security_code":   "N/A",
                "_source":      "Individual (heuristic)",
                "_master_match": "",
            }
        return {
            "largest_shareholder_company_name_en": name,
            "largest_shareholder_listing_status":  "",
            "largest_shareholder_security_code":   "",
            "_source":      "DART unmatched",
            "_master_match": "",
        }

    from modules.mapping.translation import normalize_en
    eng_name_raw = (match_row.get("corp_name_eng") or "").strip().upper() or name
    eng_name = normalize_en(eng_name_raw)
    stock    = (match_row.get("stock_code") or "").strip()
    if stock and stock.isdigit() and len(stock) == 6:
        listing = "KOSPI / KOSDAQ"
        try:
            from modules.mapping.lookup import compute_isin_from_stock_code
            isin = compute_isin_from_stock_code(stock)
        except Exception:
            isin = ""
    else:
        listing, isin = "Not listed", ""

    return {
        "largest_shareholder_company_name_en": eng_name,
        "largest_shareholder_listing_status":  listing,
        "largest_shareholder_security_code":   isin,
        "_source":      f"DART {match_kind}",
        "_master_match": match_row.get("corp_name", ""),
    }


def llm_largest_shareholder(
    items: list[dict],
    api_key: str | None,
) -> list[dict]:
    """최대주주 + 회사 영문 정의 batch.
    items: [{name_kr, name_en, mandata_brand_name, induty_code}]."""
    if not items or not api_key:
        return [{} for _ in items]
    try:
        from modules.mapping.translation import llm as _llm
        return _llm.llm_largest_shareholder_batch(items, api_key, chunk_size=20, max_workers=4)
    except Exception:
        return [{} for _ in items]


def llm_find_parents(
    items: list[dict],
    api_key: str | None,
) -> list[dict]:
    """비상장 회사들 → 모회사 추정. items: [{name_kr, name_en, stock_code, induty_code}].

    Returns: [{parent_kr, parent_en, parent_stock_code, parent_isin, status_kind, confidence}].
    """
    if not items or not api_key:
        return [{} for _ in items]
    try:
        from modules.mapping.translation import llm as _llm
        return _llm.llm_find_parents_batch(items, api_key, chunk_size=20, max_workers=4)
    except Exception:
        return [{} for _ in items]


# DART corp_cls → listing_status
CORP_CLS_TO_STATUS: dict[str, str] = {
    "Y": "KOSPI",
    "K": "KOSDAQ",
    "N": "KONEX",
    "E": "Not listed",
}


# ══════════════════════════════════════════════════════════════════════════════
# 회사 마스터 행 빌드 — DART + ISIN 정보 통합
# ══════════════════════════════════════════════════════════════════════════════

def build_company_list_row(info: dict, isin: str = "") -> dict:
    """DART company_info dict + ISIN → LIST 시트 한 행 dict.

    info 필수 키: corp_name_eng, jurir_no, induty_code, corp_cls, stock_code
    """
    eng_name = (info.get("corp_name_eng") or info.get("corp_name", "")).strip()
    jurir    = str(info.get("jurir_no", "")).strip()
    induty   = str(info.get("induty_code", "")).strip()
    corp_cls = info.get("corp_cls", "")

    # mandata_brand_name = upper(name) + "_ALL"
    mandata_brand_name = (eng_name.upper() + "_ALL") if eng_name else ""
    # mandata_brand_code = jurir_no + "01" (jurir_no 13자리 + 01 = 15자리)
    mandata_brand_code = (jurir + "01") if jurir else ""

    # GICS 매핑
    gics = map_ksic_to_gics(induty)
    if gics:
        gics_ind_code, gics_ind, gics_sub_code, gics_sub = gics
    else:
        gics_ind_code, gics_ind, gics_sub_code, gics_sub = ("", "", "", "")

    listing_status = CORP_CLS_TO_STATUS.get(corp_cls, "Not listed")

    return {
        "gics_industry_code":     gics_ind_code,
        "gics_industry":          gics_ind,
        "gics_sub_industry_code": gics_sub_code,
        "gics_sub_industry":      gics_sub,
        "company_name_en":        eng_name,
        "korea_company_number":   jurir,
        "mandata_brand_name":     mandata_brand_name,
        "mandata_brand_code":     mandata_brand_code,
        "listing_status":         listing_status,
        "security_code":          isin or "",
        # 최대주주·설명 — 기본 빈 값 (M5 에서 사용자/LLM 채움)
        "largest_shareholder_company_name_en": "",
        "largest_shareholder_listing_status":  "",
        "largest_shareholder_security_code":   "",
        "mandata_brand_name_definition":       "",
    }


# ══════════════════════════════════════════════════════════════════════════════
# 4-시트 xlsx 빌더
# ══════════════════════════════════════════════════════════════════════════════

LIST_COLUMNS = [
    "gics_industry_code", "gics_industry",
    "gics_sub_industry_code", "gics_sub_industry",
    "company_name_en", "korea_company_number",
    "mandata_brand_name", "mandata_brand_code",
    "listing_status", "security_code",
    # ── 최대주주 정보 (LLM 채움 또는 수동 편집) ──
    "largest_shareholder_company_name_en",
    "largest_shareholder_listing_status",
    "largest_shareholder_security_code",
    "mandata_brand_name_definition",
]

INFORMATION_HEADERS = [
    "Dataset Type",
    "Sector Coverage",
    "Examples of associated  Ticker",   # 원본에 공백 2개 (의도된 그대로 유지)
    "Ticker Coverage",
    "Description",
]


def build_list_rows(
    pairs: list[dict],
    match_df: "pd.DataFrame",
    info_map: dict,
    isin_map: dict,
    gics_map: dict,
    parent_overrides: dict,
    shareholder_map: dict | None = None,
    brand_to_en=None,
    product_to_en=None,
) -> tuple[list[dict], list[str]]:
    """페어(회사·브랜드·제품) → LIST 행 리스트.

    Args:
        pairs: [{company, brand, product}, ...] (M1 입력)
        match_df: DART 매칭 결과
        info_map: {corp_code: 회사상세}
        isin_map: {company_kr: isin}
        gics_map: {company_kr: gics 4-tuple dict}
        parent_overrides: {company_kr: parent info}
        shareholder_map: {company_kr: largest_shareholder_* fields}
        brand_to_en, product_to_en: 한글→영문 변환 callable. None 이면 한글 그대로

    Returns: (list_rows, pair_companies) — 같은 길이 리스트.
        pair_companies[i] = 그 행에 해당하는 회사 한글명 (검증 lookup 용)
    """
    from collections import defaultdict
    from modules.mapping.translation import normalize_en

    if brand_to_en is None:
        brand_to_en = lambda b: ("ALL" if (not b or b in ("전체","ALL","All","all")) else b)
    if product_to_en is None:
        product_to_en = lambda p: ("ALL" if (not p or p in ("전체","ALL","All","all")) else p)

    shareholder_map = shareholder_map or {}
    list_rows: list[dict] = []
    pair_companies: list[str] = []
    seq_by_company: dict[str, int] = defaultdict(int)

    for pair in pairs:
        inp        = pair.get("company", "")
        brand_kr   = pair.get("brand", "")
        product_kr = pair.get("product", "")
        m  = match_df[match_df["input_name"] == inp]
        if m.empty:
            continue
        r  = m.iloc[0]
        cc = r["corp_code"]
        info = info_map.get(cc) or {}
        info_for_row = {
            **info,
            "corp_name_eng": r["corp_name_eng"] or info.get("corp_name_eng", ""),
            "jurir_no":      info.get("jurir_no", ""),
            "stock_code":    r["stock_code"] or info.get("stock_code", ""),
            "induty_code":   info.get("induty_code", ""),
            "corp_cls":      info.get("corp_cls", ""),
        }
        row = build_company_list_row(info_for_row, isin=isin_map.get(inp, ""))

        # GICS override (사용자 검수값)
        g = gics_map.get(inp) or {}
        if g.get("gics_industry_code"):
            row["gics_industry_code"]     = g["gics_industry_code"]
            row["gics_industry"]          = g["gics_industry"]
            row["gics_sub_industry_code"] = g["gics_sub_industry_code"]
            row["gics_sub_industry"]      = g["gics_sub_industry"]
        # 비상장 → 모회사 override
        p = parent_overrides.get(inp)
        if p and p.get("parent_kr"):
            row = apply_subsidiary_to_row(row, p)
        # 최대주주·정의 (회사 단위)
        sh = shareholder_map.get(inp) or {}
        for k in ("largest_shareholder_company_name_en",
                  "largest_shareholder_listing_status",
                  "largest_shareholder_security_code",
                  "mandata_brand_name_definition"):
            if sh.get(k):
                row[k] = sh[k]

        # mandata_brand_name = COMPANY_BRAND[_PRODUCT]
        company_en = (row.get("company_name_en") or inp).upper()
        brand_en   = brand_to_en(brand_kr).upper()
        parts      = [company_en, brand_en]
        if product_kr:
            parts.append(product_to_en(product_kr).upper())
        row["mandata_brand_name"] = normalize_en(" ".join(parts))

        # mandata_brand_code = jurir + 회사단위 3자리 순번
        jurir = info.get("jurir_no", "")
        seq_by_company[inp] += 1
        row["mandata_brand_code"] = f"{jurir}{seq_by_company[inp]:03d}" if jurir else ""

        # company_name_en 도 정규화 한 번 더 (DART 의 'CO.,LTD' 같은 leftover 정리)
        row["company_name_en"] = normalize_en(row.get("company_name_en") or "")

        list_rows.append(row)
        pair_companies.append(inp)

    return list_rows, pair_companies


def build_list_only_xlsx(list_df: pd.DataFrame) -> bytes:
    """LIST_PR 단일 시트 xlsx — 회사 마스터 빌더의 표준 출력.

    INFORMATION 시트와 BT 시트는 제외. 10개 LIST 컬럼만 포함.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "LIST_PR"

    # 헤더
    for col_idx, h in enumerate(LIST_COLUMNS, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DBEAFE")

    # 데이터
    for r_idx, (_, row) in enumerate(list_df.iterrows(), start=2):
        for c_idx, col in enumerate(LIST_COLUMNS, start=1):
            v = row.get(col, "")
            if pd.isna(v):
                v = ""
            ws.cell(row=r_idx, column=c_idx, value=v)

    # 컬럼 너비
    widths = [16, 22, 18, 30, 30, 22, 38, 22, 30, 18,
              38, 32, 22, 60]   # 추가 4 컬럼
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # wrap_text for 마지막 정의 컬럼
    from openpyxl.styles import Alignment
    for r_idx in range(2, ws.max_row + 1):
        ws.cell(row=r_idx, column=len(LIST_COLUMNS)).alignment = Alignment(
            wrap_text=True, vertical="top"
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pos_master_xlsx(
    list_df: pd.DataFrame,
    info_meta: dict,
    bt_overrides: dict | None = None,
) -> bytes:
    """4-시트 xlsx bytes 생성.

    Args:
        list_df: LIST_PR 컬럼 10개를 가진 DataFrame
        info_meta: INFORMATION_PR 메타 데이터
            { "dataset_type": str, "sector_coverage": str,
              "examples_ticker": str (multi-line),
              "ticker_coverage": str, "description": str }
        bt_overrides: BT 시트(베타) 만 따로 채우고 싶을 때 dict (없으면 PR 과 동일)

    Returns: xlsx bytes
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    # 기본 시트 제거 후 순서대로 추가
    wb.remove(wb.active)

    def _make_information_sheet(name: str, meta: dict):
        ws = wb.create_sheet(name)
        # 타이틀 셀 (행 3, 열 C)
        ws.cell(row=3, column=3, value="[POS Data Introduction]")
        ws.cell(row=3, column=3).font = Font(bold=True, size=14)
        # 헤더 행 (행 5): 컬럼 C, D, E, F, H (원본 레이아웃과 동일)
        header_cols = [3, 4, 5, 6, 8]   # C, D, E, F, H
        for col_idx, header in zip(header_cols, INFORMATION_HEADERS):
            c = ws.cell(row=5, column=col_idx, value=header)
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="E5E7EB")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # 데이터 행 (행 6)
        values = [
            meta.get("dataset_type", ""),
            meta.get("sector_coverage", ""),
            meta.get("examples_ticker", ""),
            meta.get("ticker_coverage", ""),
            meta.get("description", ""),
        ]
        for col_idx, v in zip(header_cols, values):
            c = ws.cell(row=6, column=col_idx, value=v)
            c.alignment = Alignment(wrap_text=True, vertical="top")
        # 컬럼 너비
        for col_idx in header_cols:
            letter = ws.cell(row=5, column=col_idx).column_letter
            ws.column_dimensions[letter].width = 28
        ws.row_dimensions[6].height = 120

    def _make_list_sheet(name: str, df: pd.DataFrame):
        ws = wb.create_sheet(name)
        for col_idx, h in enumerate(LIST_COLUMNS, start=1):
            c = ws.cell(row=1, column=col_idx, value=h)
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="DBEAFE")
        for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
            for c_idx, col in enumerate(LIST_COLUMNS, start=1):
                v = row.get(col, "")
                # 빈 값 / NaN 처리
                if pd.isna(v):
                    v = ""
                ws.cell(row=r_idx, column=c_idx, value=v)
        # 컬럼 너비
        widths = [16, 22, 18, 30, 30, 22, 38, 22, 18, 18]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # 원본 시트 순서: INFORMATION_PR, INFORMATION_BT, LIST_PR, LIST_BT
    _make_information_sheet("INFORMATION_PR", info_meta)
    _make_information_sheet("INFORMATION_BT", (bt_overrides or {}).get("info_meta") or info_meta)
    _make_list_sheet("LIST_PR", list_df)
    _make_list_sheet("LIST_BT", (bt_overrides or {}).get("list_df", list_df))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
