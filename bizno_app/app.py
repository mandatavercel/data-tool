"""
사업자번호 → 업체명 자동 조회 (Streamlit 웹앱)
================================================
- 엑셀(사업자번호 1열) 업로드 또는 직접 입력
- bizno.net 스크래핑으로 업체명/대표자/주소/업종 조회
- (옵션) 국세청 API로 사업자 상태 조회
- 결과를 표로 보고 엑셀로 다운로드

포트: 8503 (.command 또는 streamlit run app.py --server.port 8503)
"""
from __future__ import annotations

import io
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime

import pandas as pd
import requests
import streamlit as st
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# 영문 머천트 코드 → 한글명 매핑 모듈 (Claude API)
import sys as _sys
from pathlib import Path as _Path
_sys.path.insert(0, str(_Path(__file__).resolve().parent))
import merchant_mapper  # noqa: E402


# ─────────────────────────────────────────────────────────────
# 페이지 설정
# ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="업체정보 매핑 도구",
    page_icon="🏢",
    layout="wide",
)


# ─────────────────────────────────────────────────────────────
# 모드 2: 영문 머천트 코드 → 한글 업체명 매핑 (Claude API)
# ─────────────────────────────────────────────────────────────
def render_merchant_translation_mode():
    """영문 머천트 코드 → 한글 업체명 매핑 UI."""
    st.title("🌐 영문 머천트 코드 → 한글 업체명")
    st.caption("Claude API로 영문 코드(`COUPANGEATS_MUGPOS__COMBINED` 등)를 한글 업체명/카테고리/그룹으로 자동 매핑.")

    with st.sidebar:
        st.subheader("⚙️ 설정 (영문 매핑)")
        api_key = st.text_input(
            "Anthropic API Key",
            value=st.session_state.get("anthropic_api_key", ""),
            type="password",
            help="console.anthropic.com에서 발급. sk-ant- 로 시작",
            key="anthropic_api_key_input",
        )
        if api_key:
            st.session_state["anthropic_api_key"] = api_key

        use_cache = st.checkbox(
            "캐시 사용 (이미 매핑된 코드 재사용)",
            value=True,
            help="bizno_app/data/merchant_cache.json 에 결과 저장. 다음번에 같은 코드 다시 부르면 API 호출 안 함",
        )

        # 캐시 통계
        stats = merchant_mapper.get_cache_stats()
        st.caption(f"💾 캐시: **{stats['count']:,}개** 코드 저장됨")
        if stats["count"] > 0:
            if st.button("🗑 캐시 비우기", use_container_width=True):
                merchant_mapper.clear_cache()
                st.rerun()
        st.divider()
        st.caption("결과는 새로고침하면 사라지니, 작업 후 엑셀로 꼭 다운로드하세요.")

    # ─── 입력 ──────────────────────────────────────────────────────
    up = st.file_uploader(
        "CSV/엑셀 업로드 (영문 머천트 코드가 있는 열을 선택)",
        type=["xlsx", "xls", "csv"],
        key="merchant_upload",
    )

    df_in = None
    code_col = None
    if up is not None:
        try:
            if up.name.lower().endswith(".csv"):
                df_in = pd.read_csv(up, dtype=str, keep_default_na=False)
            else:
                df_in = pd.read_excel(up, dtype=str)
            st.write(f"행 수: **{len(df_in):,}**  /  컬럼: {list(df_in.columns)}")
            # 영문/언더스코어/대문자 비중 높은 열을 디폴트로 추천
            def _en_score(s):
                if not isinstance(s, str): return 0
                return sum(1 for c in s if c.isupper() or c == "_") / max(len(s), 1)
            scores = [(c, df_in[c].dropna().head(20).map(_en_score).mean() if len(df_in) else 0) for c in df_in.columns]
            scores.sort(key=lambda x: -x[1])
            default_col = scores[0][0] if scores else df_in.columns[0]
            code_col = st.selectbox(
                "영문 머천트 코드 컬럼",
                df_in.columns.tolist(),
                index=df_in.columns.tolist().index(default_col),
            )
        except Exception as e:
            st.error(f"파일 읽기 실패: {e}")

    if df_in is None or code_col is None:
        st.info("👆 파일을 업로드하고 영문 코드가 들어있는 열을 선택하세요.")
        st.markdown(
            """
            **이 모드는 이런 데이터에 적합해요:**
            - 카드 결제 데이터에 가맹점이 영문 코드(`COUPANGEATS_MUGPOS__COMBINED`)로 들어있을 때
            - 사업자번호가 손실됐거나 없을 때, 영문 브랜드 코드만으로 회사를 식별해야 할 때
            - PG/카드사 raw 데이터의 merchant_name 컬럼

            **결과는 이렇게 나와요:**

            | code | 한글명 | 카테고리 | 그룹 | 신뢰도 |
            |---|---|---|---|---|
            | COUPANGEATS_MUGPOS__COMBINED | 쿠팡이츠 | 배달/음식 | 쿠팡 | 0.95 |
            | TOSSPAYMENTS__SINGLE | 토스페이먼츠 | 결제/PG | 비바리퍼블리카 | 0.95 |
            """
        )
        return

    # 유니크 코드 추출
    raw_codes = df_in[code_col].astype(str).str.strip().tolist()
    unique_codes = sorted(set(c for c in raw_codes if c))
    cache = merchant_mapper.load_cache() if use_cache else {}
    cached_n = sum(1 for c in unique_codes if c in cache)
    to_call_n = len(unique_codes) - cached_n

    c1, c2, c3 = st.columns(3)
    c1.metric("전체 행", f"{len(raw_codes):,}")
    c2.metric("유니크 코드", f"{len(unique_codes):,}")
    c3.metric("API 호출 필요", f"{to_call_n:,}", delta=f"-{cached_n} 캐시 적중", delta_color="off")

    # ─── 실행 ──────────────────────────────────────────────────────
    can_run = bool(api_key) and len(unique_codes) > 0
    run = st.button(
        f"🚀 Claude로 매핑 시작 ({to_call_n:,}개 API 호출)",
        type="primary",
        disabled=not can_run,
    )
    if not api_key:
        st.warning("👈 사이드바에 Anthropic API Key를 넣어주세요. (sk-ant- 로 시작)")

    if run:
        progress = st.progress(0.0)
        status = st.empty()
        status.info(f"🤖 Claude 호출 중... ({to_call_n:,}개, 배치당 30개)")

        def _cb(done, total):
            if total > 0:
                progress.progress(done / total)
                status.info(f"🤖 진행 {done}/{total}")

        try:
            results = merchant_mapper.translate_codes(
                unique_codes,
                api_key=api_key,
                use_cache=use_cache,
                progress_cb=_cb,
            )
        except RuntimeError as e:
            st.error(str(e))
            return
        except Exception as e:
            st.error(f"API 호출 실패: {e}")
            return

        progress.progress(1.0)
        status.success(f"✅ 매핑 완료! ({len(results):,}개)")

        # 결과 표 만들기 (입력 순서 유지)
        rows_out = []
        for i, code in enumerate(raw_codes, 1):
            r = results.get(code, {})
            rows_out.append({
                "No.": i,
                "원본코드": code,
                "한글명": r.get("korean_name", ""),
                "카테고리": r.get("category", ""),
                "그룹": r.get("group", ""),
                "신뢰도": round(r.get("confidence", 0.0), 2),
                "캐시": "✓" if r.get("cached") else "",
            })
        # 입력 dataframe의 다른 컬럼들도 함께 보존
        df_out = pd.DataFrame(rows_out)
        other_cols = [c for c in df_in.columns if c != code_col]
        if other_cols:
            df_out = pd.concat([df_out, df_in[other_cols].reset_index(drop=True)], axis=1)

        st.session_state["merchant_result_df"] = df_out

    # ─── 결과 표시 + 편집 ──────────────────────────────────────────
    if "merchant_result_df" in st.session_state:
        df_out = st.session_state["merchant_result_df"]

        st.subheader("📋 매핑 결과 (직접 수정 가능)")
        st.caption("표를 클릭해서 한글명/카테고리/그룹을 직접 수정할 수 있어요. 수정 후 다운로드하면 반영됨.")

        # 신뢰도 낮은 행 강조 위해 필터
        f1, f2 = st.columns([1, 3])
        with f1:
            min_conf = st.slider("신뢰도 ≥", 0.0, 1.0, 0.0, 0.1)
        df_view = df_out[df_out["신뢰도"] >= min_conf].copy()

        edited = st.data_editor(
            df_view,
            use_container_width=True,
            hide_index=True,
            num_rows="fixed",
            column_config={
                "신뢰도": st.column_config.NumberColumn(format="%.2f", disabled=True),
                "캐시": st.column_config.TextColumn(disabled=True),
                "원본코드": st.column_config.TextColumn(disabled=True),
                "No.": st.column_config.NumberColumn(disabled=True),
            },
            key="merchant_editor",
        )

        # 수정사항 반영
        if edited is not None and not edited.equals(df_view):
            for idx in edited.index:
                df_out.loc[idx, ["한글명", "카테고리", "그룹"]] = edited.loc[idx, ["한글명", "카테고리", "그룹"]]
            st.session_state["merchant_result_df"] = df_out
            st.toast("수정사항이 반영됐어요. 다운로드에 포함됩니다.", icon="✅")

        # ─── 통계 ─────────────────────────────────────────────────
        mapped_n = int((df_out["한글명"] != "").sum())
        low_conf_n = int((df_out["신뢰도"] < 0.7).sum())
        s1, s2, s3, s4 = st.columns(4)
        s1.metric("전체", f"{len(df_out):,}")
        s2.metric("매핑됨", f"{mapped_n:,}")
        s3.metric("신뢰도 < 0.7 (확인필요)", f"{low_conf_n:,}")
        s4.metric("카테고리 종류", df_out["카테고리"].nunique())

        # ─── 다운로드 ─────────────────────────────────────────────
        fname_base = f"머천트매핑_{datetime.now():%Y%m%d_%H%M}"
        d1, d2 = st.columns(2)
        with d1:
            st.download_button(
                "💾 엑셀로 다운로드",
                data=_build_simple_xlsx(df_out),
                file_name=f"{fname_base}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )
        with d2:
            st.download_button(
                "📄 CSV로 다운로드",
                data=df_out.to_csv(index=False).encode("utf-8-sig"),
                file_name=f"{fname_base}.csv",
                mime="text/csv",
                use_container_width=True,
            )


def _build_simple_xlsx(df: pd.DataFrame) -> bytes:
    """머천트 매핑용 간단한 엑셀 생성."""
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as w:
        df.to_excel(w, sheet_name="매핑결과", index=False)
        # 컬럼 너비
        ws = w.sheets["매핑결과"]
        widths = {"원본코드": 35, "한글명": 20, "카테고리": 14, "그룹": 14, "신뢰도": 8, "캐시": 6}
        for col_idx, name in enumerate(df.columns, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(name, 16)
    return bio.getvalue()


# ─────────────────────────────────────────────────────────────
# 모드 3: 마스터 파일 매칭 (범용 VLOOKUP)
# ─────────────────────────────────────────────────────────────
def _load_table(uploaded, key_label: str):
    """업로드된 파일을 DataFrame으로 로드. 첫 시트 사용. dtype=str."""
    name = uploaded.name.lower()
    if name.endswith(".csv"):
        df = pd.read_csv(uploaded, dtype=str, keep_default_na=False)
    elif name.endswith((".xlsx", ".xls")):
        df = pd.read_excel(uploaded, dtype=str)
        df = df.fillna("")
    else:
        raise ValueError(f"지원되지 않는 파일 형식: {name}")
    return df


def _normalize_key(series: pd.Series, *, trim: bool, lower: bool, remove_separators: bool) -> pd.Series:
    """매칭용 키를 정규화. 원본은 보존."""
    s = series.astype(str)
    if trim:
        s = s.str.strip()
    if lower:
        s = s.str.lower()
    if remove_separators:
        # 공백, -, _, . 제거
        s = s.str.replace(r"[\s\-_.]", "", regex=True)
    return s


def render_master_match_mode():
    """두 파일을 기준 컬럼으로 매칭해서 마스터의 정보를 가져오는 범용 VLOOKUP."""
    st.title("🔗 마스터 파일 매칭")
    st.caption("두 파일에서 기준 컬럼을 골라 매칭하고, 마스터의 사업자번호(또는 기타) 컬럼을 가져옵니다. VLOOKUP/INDEX-MATCH의 일반화 버전.")

    with st.sidebar:
        st.subheader("⚙️ 매칭 옵션")
        opt_trim = st.checkbox("앞뒤 공백 제거", value=True)
        opt_lower = st.checkbox("대소문자 무시", value=True)
        opt_remove_sep = st.checkbox(
            "구분자 제거 (공백/-/_/.)",
            value=False,
            help="키 비교 전에 공백·하이픈·언더스코어·점을 모두 제거. 예: 'COUPANG EATS' == 'COUPANGEATS'",
        )
        opt_keep_first = st.checkbox(
            "마스터 중복 시 첫 번째만 사용",
            value=True,
            help="마스터 파일에서 같은 키가 여러 번 나오면 첫 번째 행의 값을 가져옴",
        )
        st.divider()
        st.caption("결과는 새로고침하면 사라지니, 작업 후 꼭 다운로드하세요.")

    col_left, col_right = st.columns(2, gap="large")

    # ─── 메인 파일 ────────────────────────────────────────────
    with col_left:
        st.markdown("### 📥 메인 파일 (대상)")
        st.caption("이 파일에 마스터의 값을 붙여넣을 거예요.")
        up_main = st.file_uploader(
            "메인 파일 업로드",
            type=["xlsx", "xls", "csv"],
            key="match_main_upload",
        )
        df_main = None
        key_main = None
        if up_main is not None:
            try:
                df_main = _load_table(up_main, "메인")
                st.caption(f"행 {len(df_main):,}  ·  컬럼 {len(df_main.columns)}")
                key_main = st.selectbox(
                    "메인 파일의 **기준 컬럼**",
                    df_main.columns.tolist(),
                    key="match_main_key",
                )
                with st.expander("미리보기 (앞 5행)"):
                    st.dataframe(df_main.head(5), use_container_width=True, hide_index=True)
            except Exception as e:
                st.error(f"파일 읽기 실패: {e}")

    # ─── 마스터 파일 ────────────────────────────────────────────
    with col_right:
        st.markdown("### 📚 마스터 파일 (참조)")
        st.caption("여기서 값을 가져올 거예요.")
        up_master = st.file_uploader(
            "마스터 파일 업로드",
            type=["xlsx", "xls", "csv"],
            key="match_master_upload",
        )
        df_master = None
        key_master = None
        bring_cols = []
        if up_master is not None:
            try:
                df_master = _load_table(up_master, "마스터")
                st.caption(f"행 {len(df_master):,}  ·  컬럼 {len(df_master.columns)}")
                key_master = st.selectbox(
                    "마스터 파일의 **기준 컬럼**",
                    df_master.columns.tolist(),
                    key="match_master_key",
                )
                # 가져올 컬럼들 (기준 컬럼 제외 디폴트 추천: '사업자번호' 비슷한 이름)
                non_key = [c for c in df_master.columns if c != key_master]
                # 디폴트: 사업자/biz/number 들어간 컬럼은 자동 선택
                def _looks_biz(c: str) -> bool:
                    lc = c.lower()
                    return any(k in lc for k in ["사업자", "biz", "사업자번호", "사업자등록", "번호", "no.", "registration"])
                default_pick = [c for c in non_key if _looks_biz(c)]
                if not default_pick:
                    default_pick = non_key[:1]  # 적어도 한 개
                bring_cols = st.multiselect(
                    "**가져올 컬럼들** (다중 선택 가능)",
                    options=non_key,
                    default=default_pick,
                    key="match_bring_cols",
                )
                with st.expander("미리보기 (앞 5행)"):
                    st.dataframe(df_master.head(5), use_container_width=True, hide_index=True)
            except Exception as e:
                st.error(f"파일 읽기 실패: {e}")

    # ─── 실행 ───────────────────────────────────────────────────
    if df_main is None or df_master is None:
        st.info("👆 메인 파일과 마스터 파일을 모두 업로드하세요.")
        return
    if not key_main or not key_master or not bring_cols:
        st.warning("기준 컬럼과 가져올 컬럼을 선택하세요.")
        return

    st.divider()
    run = st.button("🚀 매칭 실행", type="primary", use_container_width=False)

    if run or "match_result_df" in st.session_state:
        if run:
            # 정규화 키 생성
            main_key_norm = _normalize_key(
                df_main[key_main],
                trim=opt_trim, lower=opt_lower, remove_separators=opt_remove_sep,
            )
            master_key_norm = _normalize_key(
                df_master[key_master],
                trim=opt_trim, lower=opt_lower, remove_separators=opt_remove_sep,
            )

            # 마스터에서 가져올 부분만 떼고 정규화 키 붙임
            master_pick = df_master[[key_master, *bring_cols]].copy()
            master_pick["_norm_key_"] = master_key_norm

            # 중복 처리
            dup_mask = master_pick["_norm_key_"].duplicated(keep=False)
            n_dup_keys = master_pick.loc[dup_mask, "_norm_key_"].nunique()
            if opt_keep_first:
                master_pick = master_pick.drop_duplicates(subset="_norm_key_", keep="first")

            # 컬럼명 충돌 회피
            renamed_bring = {}
            for c in bring_cols:
                new_name = c
                while new_name in df_main.columns:
                    new_name = f"{new_name}_(마스터)"
                renamed_bring[c] = new_name
            master_pick = master_pick.rename(columns=renamed_bring)
            also_keep_master_key = f"{key_master}_(마스터)"
            while also_keep_master_key in df_main.columns:
                also_keep_master_key += "_"
            master_pick = master_pick.rename(columns={key_master: also_keep_master_key})

            # merge
            main_join = df_main.copy()
            main_join["_norm_key_"] = main_key_norm
            merged = main_join.merge(
                master_pick,
                on="_norm_key_",
                how="left",
            )
            merged = merged.drop(columns=["_norm_key_"])

            # 매칭 여부 계산
            check_col = list(renamed_bring.values())[0] if renamed_bring else also_keep_master_key
            matched_mask = merged[check_col].notna() & (merged[check_col].astype(str) != "")
            merged["_매칭_"] = matched_mask.map(lambda x: "✅ 매칭" if x else "❌ 미매칭")

            st.session_state["match_result_df"] = merged
            st.session_state["match_stats"] = {
                "total": len(merged),
                "matched": int(matched_mask.sum()),
                "unmatched": int((~matched_mask).sum()),
                "master_total": len(df_master),
                "master_dup_keys": int(n_dup_keys),
                "bring_cols": list(renamed_bring.values()),
                "master_key_col": also_keep_master_key,
            }

        merged = st.session_state["match_result_df"]
        stats = st.session_state["match_stats"]

        # ─── 결과 요약 ───────────────────────────────────────
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("메인 전체", f"{stats['total']:,}")
        c2.metric("✅ 매칭됨", f"{stats['matched']:,}")
        c3.metric("❌ 미매칭", f"{stats['unmatched']:,}")
        match_rate = stats["matched"] / max(stats["total"], 1) * 100
        c4.metric("매칭률", f"{match_rate:.1f}%")
        if stats["master_dup_keys"] > 0:
            st.warning(
                f"⚠️ 마스터 파일에 같은 키가 중복되는 경우 **{stats['master_dup_keys']}개 키**가 있어요. "
                f"{'첫 번째만 사용했어요.' if opt_keep_first else '모든 중복이 그대로 조인돼서 행 수가 늘어났을 수 있어요.'}"
            )

        # ─── 결과 표 ─────────────────────────────────────────
        st.subheader("📋 매칭 결과")
        view_filter = st.radio(
            "표시",
            ["전체", "매칭된 것만", "미매칭만"],
            horizontal=True,
            key="match_filter",
        )
        if view_filter == "매칭된 것만":
            df_view = merged[merged["_매칭_"] == "✅ 매칭"]
        elif view_filter == "미매칭만":
            df_view = merged[merged["_매칭_"] == "❌ 미매칭"]
        else:
            df_view = merged

        st.dataframe(df_view, use_container_width=True, hide_index=True)

        # ─── 다운로드 ────────────────────────────────────────
        fname_base = f"매칭결과_{datetime.now():%Y%m%d_%H%M}"
        d1, d2, d3 = st.columns(3)
        with d1:
            st.download_button(
                "💾 전체 엑셀",
                data=_build_simple_xlsx(merged),
                file_name=f"{fname_base}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )
        with d2:
            st.download_button(
                "📄 전체 CSV",
                data=merged.to_csv(index=False).encode("utf-8-sig"),
                file_name=f"{fname_base}.csv",
                mime="text/csv",
                use_container_width=True,
            )
        with d3:
            df_unmatched = merged[merged["_매칭_"] == "❌ 미매칭"]
            if len(df_unmatched) > 0:
                st.download_button(
                    f"⚠️ 미매칭만 ({len(df_unmatched):,})",
                    data=df_unmatched.to_csv(index=False).encode("utf-8-sig"),
                    file_name=f"{fname_base}_미매칭.csv",
                    mime="text/csv",
                    use_container_width=True,
                    help="다음 단계로 처리해야 하는 행 목록",
                )


# ─────────────────────────────────────────────────────────────
# 상수 / 유틸
# ─────────────────────────────────────────────────────────────
HEADER_BG = "1F3864"
NTS_API_KEY_DEFAULT = "7ae65fcc8dc5a72a91715806d242d685c31432d2fd64b6f03a7fc347f4c15772"
NTS_API_URL = "https://api.odcloud.kr/api/nts-businessman/v1/status"
NTS_BATCH_SIZE = 100

UA = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)


def normalize_bizno(raw) -> str | None:
    """입력값을 10자리 사업자번호 문자열로 정규화. 실패 시 None."""
    if raw is None:
        return None
    if isinstance(raw, float):
        try:
            raw = str(int(raw))
        except Exception:
            return None
    s = re.sub(r"\D", "", str(raw).strip())
    if not s:
        return None
    s = s.zfill(10)
    return s if len(s) == 10 else None


def format_bizno(s: str) -> str:
    return f"{s[:3]}-{s[3:5]}-{s[5:]}"


# ─────────────────────────────────────────────────────────────
# 입력 코드 분류기 — Excel 정밀도 손실, 점포코드 등 자동 감지
# ─────────────────────────────────────────────────────────────
SCI_PATTERN = re.compile(r"^\s*-?\d+(\.\d+)?[Ee][+\-]?\d+\s*$")


def classify_code(raw, use_first10_for_long: bool = False) -> dict:
    """
    한 셀의 원본값을 분류.
    반환:
      {
        'original': 원본 문자열,
        'kind':  'valid' | 'lossy' | 'nonstandard' | 'invalid',
        'biz_no': '1234567890' or None,   # kind='valid'일 때만 채워짐
        'reason': 분류 사유 (사용자에게 보여줄 한국어 설명),
      }
    """
    out = {"original": "", "kind": "invalid", "biz_no": None, "reason": "빈 값"}

    if raw is None:
        return out

    if isinstance(raw, float):
        # pandas가 자동으로 float으로 읽었을 때
        try:
            if raw != raw:  # NaN
                return out
            raw_str = str(int(raw)) if raw.is_integer() else f"{raw:.0f}"
        except Exception:
            raw_str = str(raw)
    else:
        raw_str = str(raw).strip()

    out["original"] = raw_str
    if not raw_str:
        return out

    # 1) 과학적 표기법 — Excel이 CSV로 저장하며 정밀도를 날린 케이스
    if SCI_PATTERN.match(raw_str):
        out["kind"] = "lossy"
        out["reason"] = "Excel이 CSV 저장 시 정밀도 손실 (E+ 표기). 원본 엑셀에서 텍스트로 다시 export 필요"
        return out

    # 2) 숫자만 남기기
    digits = re.sub(r"\D", "", raw_str)
    if not digits:
        out["reason"] = "숫자 없음"
        return out

    # 3) 길이별 처리
    if len(digits) == 10:
        out["kind"] = "valid"
        out["biz_no"] = digits
        out["reason"] = "정상 10자리 사업자번호"
        return out

    if len(digits) < 10:
        # 짧으면 0 패딩하지만 보통은 실수 입력. 일단 valid로 처리하되 짧다고 표시
        padded = digits.zfill(10)
        out["kind"] = "valid"
        out["biz_no"] = padded
        out["reason"] = f"{len(digits)}자리 → 앞에 0 채워 사용"
        return out

    # len > 10 — 점포코드 붙은 비표준 형식 (예: 11자리, 13자리 우리형식 코드)
    if use_first10_for_long:
        out["kind"] = "valid"
        out["biz_no"] = digits[:10]
        out["reason"] = f"{len(digits)}자리 코드 → 앞 10자리만 사용 ({digits[:10]})"
    else:
        out["kind"] = "nonstandard"
        out["reason"] = f"{len(digits)}자리 비표준 코드 (사업자번호+점포코드?). '앞 10자리 사용' 옵션 켜면 조회 가능"
    return out


# ─────────────────────────────────────────────────────────────
# bizno.net 스크래퍼 (홈택스_업체명_자동조회.py 로직 이식)
# ─────────────────────────────────────────────────────────────
def query_bizno_one(biz_no: str, timeout: int = 10) -> dict:
    """bizno.net에서 한 건 조회."""
    out = {"biz_no": biz_no, "biz_name": "", "rep_name": "", "address": "", "biz_type": ""}
    try:
        url = f"https://bizno.net/article/{biz_no}"
        resp = requests.get(url, headers={"User-Agent": UA}, timeout=timeout)
        soup = BeautifulSoup(resp.text, "html.parser")

        # 상호명: h1~h3 중 의미 있는 것
        for tag in ["h1", "h2", "h3"]:
            el = soup.select_one(tag)
            if el:
                t = el.text.strip()
                if len(t) > 1 and "bizno" not in t.lower():
                    out["biz_name"] = t
                    break

        # 테이블/리스트 파싱
        for row in soup.select("table tr, dl, .info_list li"):
            ths = row.select("th")
            tds = row.select("td")
            if ths and tds:
                key = ths[0].text.strip()
                val = tds[0].text.strip()
                if "대표" in key and not out["rep_name"]:
                    out["rep_name"] = val
                elif "주소" in key and not out["address"]:
                    out["address"] = val
                elif ("업종" in key or "업태" in key) and not out["biz_type"]:
                    out["biz_type"] = val
                elif ("상호" in key or "회사" in key) and not out["biz_name"]:
                    out["biz_name"] = val
        return out
    except Exception:
        return out


def scrape_company_names(biz_numbers: list[str], workers: int, progress_cb=None) -> dict[str, dict]:
    """병렬로 업체명 조회. progress_cb(done, total, latest_no, latest_name)"""
    results: dict[str, dict] = {}
    total = len(biz_numbers)
    done = 0
    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = {pool.submit(query_bizno_one, b): b for b in biz_numbers}
        for fut in as_completed(futures):
            r = fut.result()
            results[r["biz_no"]] = r
            done += 1
            if progress_cb:
                progress_cb(done, total, r["biz_no"], r["biz_name"])
    return results


# ─────────────────────────────────────────────────────────────
# 국세청 사업자 상태 API (국세청_상태조회_실행코드.py 로직 이식)
# ─────────────────────────────────────────────────────────────
def check_nts_status(biz_numbers: list[str], api_key: str, progress_cb=None) -> dict[str, dict]:
    """국세청 API로 사업자 상태 조회. 100건씩 배치 호출."""
    results: dict[str, dict] = {}
    total = len(biz_numbers)
    if not api_key:
        return results

    url = f"{NTS_API_URL}?serviceKey={api_key}"
    headers = {"Content-Type": "application/json"}

    for i in range(0, total, NTS_BATCH_SIZE):
        batch = biz_numbers[i : i + NTS_BATCH_SIZE]
        try:
            resp = requests.post(
                url, headers=headers, json={"b_no": batch}, timeout=30
            )
            data = resp.json()
            for item in data.get("data", []):
                b = item.get("b_no", "")
                results[b] = {
                    "b_stt": item.get("b_stt", ""),
                    "tax_type": item.get("tax_type", ""),
                    "end_dt": item.get("end_dt", ""),
                }
        except Exception:
            pass
        if progress_cb:
            progress_cb(min(i + NTS_BATCH_SIZE, total), total)
        time.sleep(0.3)
    return results


# ─────────────────────────────────────────────────────────────
# 엑셀 저장 (예쁘게 — 홈택스 스크립트의 스타일 유지)
# ─────────────────────────────────────────────────────────────
def build_excel(df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "업체명_매핑결과"

    border = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )
    cols = list(df.columns)
    total = len(df)
    named = int((df.get("업체명", pd.Series([], dtype=str)).fillna("") != "").sum()) if "업체명" in df.columns else 0

    # 타이틀 행
    last_col = get_column_letter(len(cols))
    ws.merge_cells(f"A1:{last_col}1")
    title = ws["A1"]
    title.value = (
        f"사업자번호 업체명 매핑 결과  |  총 {total:,}개  |  확인 {named:,}개"
        f"  |  {datetime.now():%Y-%m-%d %H:%M}"
    )
    title.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    title.fill = PatternFill("solid", fgColor=HEADER_BG)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # 헤더
    widths_default = {
        "No.": 6,
        "사업자번호": 16,
        "사업자번호(원본)": 14,
        "업체명": 30,
        "대표자": 14,
        "주소": 40,
        "업종": 20,
        "사업자상태": 14,
        "과세유형": 16,
        "폐업일": 12,
        "비고": 12,
    }
    for col_idx, name in enumerate(cols, 1):
        cell = ws.cell(row=2, column=col_idx, value=name)
        cell.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = widths_default.get(name, 18)
    ws.row_dimensions[2].height = 24

    # 데이터
    for i, (_, row) in enumerate(df.iterrows(), 1):
        excel_row = i + 2
        fill = "FFFFFF" if i % 2 == 1 else "F2F7FB"
        for col_idx, name in enumerate(cols, 1):
            val = row[name]
            if pd.isna(val):
                val = ""
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.font = Font(name="Arial", size=9)
            cell.fill = PatternFill("solid", fgColor=fill)
            align = "center" if name in ("No.", "사업자번호", "사업자번호(원본)", "사업자상태", "폐업일") else "left"
            cell.alignment = Alignment(horizontal=align, vertical="center")
            cell.border = border

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{last_col}{total + 2}"

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ─────────────────────────────────────────────────────────────
# UI — 모드 선택 (사이드바 최상단)
# ─────────────────────────────────────────────────────────────
with st.sidebar:
    st.subheader("📂 작업 모드")
    _mode = st.radio(
        "어떤 매핑을 하시나요?",
        [
            "🏢 사업자번호 → 업체명",
            "🌐 영문코드 → 한글명",
            "🔗 마스터 파일 매칭",
        ],
        key="bizno_app_mode",
        label_visibility="collapsed",
    )
    st.divider()

# 모드 분기
if _mode == "🌐 영문코드 → 한글명":
    render_merchant_translation_mode()
    st.stop()
elif _mode == "🔗 마스터 파일 매칭":
    render_master_match_mode()
    st.stop()

# ─────────────────────────────────────────────────────────────
# 모드 1: 사업자번호 → 업체명 (기존)
# ─────────────────────────────────────────────────────────────
st.title("🏢 사업자번호 → 업체명 조회")
st.caption("bizno.net 스크래핑 + (옵션) 국세청 사업자상태 API")

with st.sidebar:
    st.subheader("⚙️ 설정")
    workers = st.slider("동시 조회 스레드 수", 1, 30, 10, help="너무 높이면 차단될 수 있어요. 보통 10~15")
    use_first10 = st.checkbox(
        "11자리+ 비표준 코드는 앞 10자리만 사용",
        value=False,
        help="우리 내부형식(사업자번호+점포코드)에서 앞 10자리를 사업자번호로 간주하고 조회. "
             "데이터 출처가 점포코드 붙은 형식임을 알 때만 켜세요.",
    )
    do_nts = st.checkbox("국세청 사업자상태도 함께 조회", value=False)
    api_key = ""
    if do_nts:
        api_key = st.text_input(
            "공공데이터포털 API Key",
            value=NTS_API_KEY_DEFAULT,
            type="password",
            help="data.go.kr 에서 '국세청_사업자등록정보 진위확인 및 상태조회' 신청",
        )
    st.divider()
    st.caption("결과는 새로고침하면 사라지니, 작업 후 엑셀로 꼭 다운로드하세요.")

# 입력 영역
tab_upload, tab_paste = st.tabs(["📂 엑셀 업로드", "✍️ 직접 입력"])

# 분류된 행 목록 (각 항목: {'original', 'kind', 'biz_no', 'reason'})
classified_rows: list[dict] = []

with tab_upload:
    up = st.file_uploader(
        "사업자번호 엑셀/CSV (1열에 사업자번호, 첫 행은 헤더 가정)",
        type=["xlsx", "xls", "csv"],
        key="bizno_upload",
    )
    if up is not None:
        try:
            if up.name.lower().endswith(".csv"):
                df_in = pd.read_csv(up, dtype=str, keep_default_na=False)
            else:
                df_in = pd.read_excel(up, dtype=str)
            st.write(f"행 수: **{len(df_in):,}**  /  컬럼: {list(df_in.columns)}")
            col_pick = st.selectbox("사업자번호가 들어있는 열", df_in.columns.tolist(), index=0)
            raw_values = df_in[col_pick].tolist()
            classified_rows = [classify_code(v, use_first10_for_long=use_first10) for v in raw_values]
        except Exception as e:
            st.error(f"엑셀 읽기 실패: {e}")

with tab_paste:
    txt = st.text_area(
        "사업자번호를 줄바꿈/쉼표로 구분해서 붙여넣기",
        height=180,
        placeholder="123-45-67890\n2233344455\n...",
        key="bizno_paste",
    )
    if txt.strip() and not classified_rows:
        pieces = [p for p in re.split(r"[\s,;]+", txt) if p]
        classified_rows = [classify_code(p, use_first10_for_long=use_first10) for p in pieces]

# ─────────────────────────────────────────────────────────────
# 분류 결과 요약 표시
# ─────────────────────────────────────────────────────────────
if classified_rows:
    cnt_valid = sum(1 for r in classified_rows if r["kind"] == "valid")
    cnt_lossy = sum(1 for r in classified_rows if r["kind"] == "lossy")
    cnt_nonstd = sum(1 for r in classified_rows if r["kind"] == "nonstandard")
    cnt_invalid = sum(1 for r in classified_rows if r["kind"] == "invalid")

    st.subheader("🔍 입력 분류 결과")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("✅ 유효 (조회대상)", f"{cnt_valid:,}")
    c2.metric("⚠️ 손실 (E+ 표기)", f"{cnt_lossy:,}", help="Excel이 CSV로 저장하며 정밀도가 깨진 행")
    c3.metric("🔶 비표준 (11+자리)", f"{cnt_nonstd:,}", help="점포코드 붙은 우리형식. 좌측 옵션으로 처리 가능")
    c4.metric("❌ 빈 값/오류", f"{cnt_invalid:,}")

    if cnt_lossy > 0:
        st.warning(
            f"⚠️  **{cnt_lossy:,}개 행이 과학적 표기법(E+)** 으로 들어있어요. "
            "Excel이 CSV로 저장할 때 13자리 숫자의 정밀도를 5~6자리로 잘라버린 상태라 **원본 복원이 불가능**합니다. "
            "원본 엑셀 파일을 열어서 해당 열을 **'텍스트' 형식으로 바꾼 뒤 다시 CSV로 저장**해 주세요. "
            "(아래 도움말 펼치기 참고)"
        )
    if cnt_nonstd > 0 and not use_first10:
        st.info(
            f"🔶  **{cnt_nonstd:,}개 행이 11자리 이상 비표준 코드**예요. "
            "왼쪽 사이드바의 **「11자리+ 비표준 코드는 앞 10자리만 사용」** 옵션을 켜면 조회 시도가 가능합니다."
        )

    with st.expander("📋 분류 상세 (앞 20개 행)"):
        preview_df = pd.DataFrame([
            {
                "원본": r["original"][:30] if r["original"] else "(빈 값)",
                "분류": {"valid": "✅ 유효", "lossy": "⚠️ 손실", "nonstandard": "🔶 비표준", "invalid": "❌ 오류"}[r["kind"]],
                "정규화": r["biz_no"] or "-",
                "사유": r["reason"],
            } for r in classified_rows[:20]
        ])
        st.dataframe(preview_df, use_container_width=True, hide_index=True)

    with st.expander("❓ E+ 표기 / 정밀도 손실 해결 방법"):
        st.markdown(
            """
            **왜 이런 일이 생기나요?**
            Excel은 11자리 이상 숫자를 셀에 넣으면 자동으로 과학적 표기법(`6.8095E+12`)으로 보여줘요.
            그 상태에서 CSV로 저장하면 표시된 그대로 저장되어 **실제 숫자 13자리 중 5~6자리만 남고 나머지가 0으로 잘립니다.**
            예: `6809512345678` → CSV에서 `6.8095E+12` → 다시 읽으면 `6809500000000` (뒤 7자리 손실)

            **해결 방법 (앞으로 데이터를 만들 때)**
            1. 원본 데이터 소스(DB/시스템)에서 export할 때 사업자번호 컬럼을 **TEXT/문자열** 타입으로 export
            2. Excel에서 그 컬럼을 클릭 → 마우스우클릭 → **셀 서식 → 텍스트** 로 먼저 바꾸고 데이터 붙여넣기
            3. CSV가 아닌 **xlsx**로 저장하면 정밀도 유지됨 (xlsx는 셀별로 타입 보존)
            4. CSV로 꼭 저장해야 하면, 앞에 작은따옴표(`'`)를 붙여서 강제 텍스트로

            **현재 파일을 보는 법 (이미 손실된 것)**
            손실된 CSV는 **복원이 안 돼요**. Excel에서 열어서 셀 서식을 바꿔도, 잘려나간 뒷자리는 돌아오지 않습니다.
            원본을 새로 받아야 해요.
            """
        )

st.divider()

# 조회 대상(중복 제거된 valid bizno) 추출
valid_bizno_unique: list[str] = []
_seen = set()
for r in classified_rows:
    if r["kind"] == "valid" and r["biz_no"] and r["biz_no"] not in _seen:
        _seen.add(r["biz_no"])
        valid_bizno_unique.append(r["biz_no"])

# 실행 버튼
col_run, col_info = st.columns([1, 3])
with col_run:
    run = st.button(
        f"🚀 조회 시작 ({len(valid_bizno_unique):,}개)",
        type="primary",
        disabled=(len(valid_bizno_unique) == 0),
    )
with col_info:
    if valid_bizno_unique:
        est_min = max(1, int(len(valid_bizno_unique) / max(workers, 1) / 6))
        st.caption(f"예상 소요시간: 약 **{est_min}분** (네트워크에 따라 변동). 손실/비표준/오류 행은 결과표에 함께 표시되지만 외부 조회는 안 합니다.")

if run and valid_bizno_unique:
    # 진행상황 영역
    progress = st.progress(0.0)
    status_box = st.empty()
    log_box = st.empty()

    # 1) bizno.net 스크래핑
    status_box.info("🌐 bizno.net에서 업체명 조회 중...")
    recent_lines: list[str] = []

    def on_progress(done: int, total: int, biz_no: str, biz_name: str):
        progress.progress(done / total)
        recent_lines.append(
            f"[{done:>4}/{total}] {format_bizno(biz_no)} → {biz_name or '(미확인)'}"
        )
        log_box.code("\n".join(recent_lines[-8:]), language="text")

    scrape_results = scrape_company_names(valid_bizno_unique, workers=workers, progress_cb=on_progress)

    # 2) (옵션) 국세청 상태 조회
    nts_results: dict[str, dict] = {}
    if do_nts and api_key:
        status_box.info("🏛 국세청 사업자상태 조회 중...")
        progress.progress(0.0)

        def on_nts_progress(done: int, total: int):
            progress.progress(done / total)

        nts_results = check_nts_status(valid_bizno_unique, api_key=api_key, progress_cb=on_nts_progress)

    status_box.success("✅ 조회 완료!")
    progress.progress(1.0)

    # 3) DataFrame — 입력된 모든 행(분류 포함)을 결과에 합쳐줌
    KIND_LABEL = {"valid": "✅ 유효", "lossy": "⚠️ 정밀도손실", "nonstandard": "🔶 비표준", "invalid": "❌ 오류"}
    rows = []
    for i, r in enumerate(classified_rows, 1):
        kind = r["kind"]
        b = r["biz_no"]
        s = scrape_results.get(b, {}) if (kind == "valid" and b) else {}
        n = nts_results.get(b, {}) if (kind == "valid" and b) else {}
        row = {
            "No.": i,
            "원본 입력": r["original"],
            "분류": KIND_LABEL[kind],
            "사업자번호": format_bizno(b) if b else "",
            "사업자번호(원본)": b or "",
            "업체명": s.get("biz_name", ""),
            "대표자": s.get("rep_name", ""),
            "주소": s.get("address", ""),
            "업종": s.get("biz_type", ""),
        }
        if do_nts:
            row["사업자상태"] = n.get("b_stt", "")
            row["과세유형"] = n.get("tax_type", "")
            row["폐업일"] = n.get("end_dt", "")
        # 비고: 분류 사유 우선, 유효지만 미확인이면 그것도 표시
        if kind != "valid":
            row["비고"] = r["reason"]
        elif not s.get("biz_name"):
            row["비고"] = "조회 실패 (미확인)"
        else:
            row["비고"] = ""
        rows.append(row)
    df = pd.DataFrame(rows)

    # 4) 요약
    named = int(((df["분류"] == KIND_LABEL["valid"]) & (df["업체명"] != "")).sum())
    valid_n = int((df["분류"] == KIND_LABEL["valid"]).sum())
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("전체 행", f"{len(df):,}")
    c2.metric("유효 조회시도", f"{valid_n:,}")
    c3.metric("업체명 확인", f"{named:,}")
    c4.metric("미확인+스킵", f"{int(len(df) - named):,}")

    # 분류별 필터
    show_kinds = st.multiselect(
        "표시할 분류",
        options=list(KIND_LABEL.values()),
        default=list(KIND_LABEL.values()),
    )
    df_view = df[df["분류"].isin(show_kinds)]
    st.dataframe(df_view, use_container_width=True, hide_index=True)

    # 5) 다운로드 (전체 / 유효만 / 손실만)
    fname_base = f"사업자번호_업체명_매핑_{datetime.now():%Y%m%d_%H%M}"
    d1, d2, d3 = st.columns(3)
    with d1:
        st.download_button(
            "💾 전체 엑셀",
            data=build_excel(df),
            file_name=f"{fname_base}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )
    with d2:
        st.download_button(
            "📄 전체 CSV",
            data=df.to_csv(index=False).encode("utf-8-sig"),
            file_name=f"{fname_base}.csv",
            mime="text/csv",
            use_container_width=True,
        )
    with d3:
        df_lossy = df[df["분류"].isin([KIND_LABEL["lossy"], KIND_LABEL["nonstandard"], KIND_LABEL["invalid"]])]
        if len(df_lossy) > 0:
            st.download_button(
                f"⚠️ 손실/스킵만 ({len(df_lossy):,})",
                data=df_lossy.to_csv(index=False).encode("utf-8-sig"),
                file_name=f"{fname_base}_손실분.csv",
                mime="text/csv",
                use_container_width=True,
                help="원본을 다시 받아야 하는 행 목록. 데이터 담당자에게 전달하세요.",
            )

elif not classified_rows:
    st.info("👈 왼쪽 탭에서 엑셀을 업로드하거나 사업자번호를 붙여넣으세요.")
